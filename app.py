import flet as ft
import flet_datatable2 as fdt
import csv
import re
from pathlib import Path
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors as pdf_colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from database import (
    add_item as add_item_to_database,
    get_items,
    update_item,
    count_inventory_for_item,
    delete_item,
    add_inventory,
    get_inventory,
    update_inventory_record,
    delete_inventory_record
)
from config import THEME, WINDOW, save_config

# Theme (see config.json to customize)
COLOUR_TEXT = THEME["text"]
COLOUR_CARD_BG = THEME["card_bg"]
COLOUR_BORDER = THEME["border"]
COLOUR_DIVIDER = THEME["divider"]
COLOUR_ROW_SEPARATOR = THEME["row_separator"]
COLOUR_XLSX_CELL = THEME["xlsx_header_bg"]
COLOUR_PDF_HEADER_BG = THEME["pdf_header_bg"]
COLOUR_SAVE_FAIL = THEME["save_fail"]
COLOUR_SAVE_SUCCESS = THEME["save_success"]


HEX_COLOR_RE = re.compile(r"^#?[0-9A-Fa-f]{6}$")


def is_valid_hex_color(value):
    return bool(HEX_COLOR_RE.match(value.strip()))


def picked_date_str(value):
    """
    Flet's DatePicker currently returns a date one day earlier than what
    was actually clicked (a known upstream UTC conversion bug —
    https://github.com/flet-dev/flet/issues/6145). Bump it back by a day
    until that's fixed upstream.
    """
    return (value + timedelta(days=1)).strftime("%Y-%m-%d")


# -------------------------
# Functions
# -------------------------

def load_items(
    page: ft.Page,
    items_list: ft.Column,
    inventory_item: ft.Dropdown,
    inventory_list: ft.Column
):
    items_list.controls.clear()

    rows = get_items()

    if not rows:
        items_list.controls.append(
            ft.Text(
                "No items yet — add one to get started.",
                italic=True,
                color=COLOUR_TEXT
            )
        )

    for item_id, name, unit in rows:
        items_list.controls.append(
            ft.Container(
                content=ft.Row(
                    controls=[
                        ft.Text(f"{name} ({unit})", expand=True),
                        ft.IconButton(
                            icon=ft.Icons.EDIT,
                            tooltip="Edit item",
                            icon_size=18,
                            on_click=lambda e, iid=item_id, nm=name, un=unit: (
                                open_edit_item_dialog(
                                    page,
                                    iid,
                                    nm,
                                    un,
                                    items_list,
                                    inventory_item,
                                    inventory_list
                                )
                            )
                        ),
                        ft.IconButton(
                            icon=ft.Icons.DELETE,
                            tooltip="Delete item",
                            icon_size=18,
                            on_click=lambda e, iid=item_id, nm=name: (
                                open_delete_item_dialog(
                                    page,
                                    iid,
                                    nm,
                                    items_list,
                                    inventory_item,
                                    inventory_list
                                )
                            )
                        )
                    ],
                    alignment=ft.MainAxisAlignment.SPACE_BETWEEN
                ),
                padding=ft.Padding.symmetric(vertical=2, horizontal=8),
                border_radius=8,
                border=ft.Border.only(
                    bottom=ft.BorderSide(1, COLOUR_BORDER)
                )
            )
        )

    page.update()

def open_edit_item_dialog(
    page: ft.Page,
    item_id: int,
    name: str,
    unit: str,
    items_list: ft.Column,
    inventory_item: ft.Dropdown,
    inventory_list: ft.Column
):
    name_field = ft.TextField(label="Name", value=name)
    unit_field = ft.TextField(label="Unit", value=unit)
    message = ft.Text()

    def save(e):
        new_name = name_field.value.strip()
        new_unit = unit_field.value.strip()

        if not new_name or not new_unit:
            return

        if not update_item(item_id, new_name, new_unit):
            message.value = "An item with that name already exists."
            page.update()
            return

        load_items(page, items_list, inventory_item, inventory_list)
        load_inventory_items(page, inventory_item)
        load_inventory(page, inventory_list)

        page.pop_dialog()

    page.show_dialog(
        ft.AlertDialog(
            modal=True,
            title=ft.Text("Edit Item"),
            content=ft.Column(
                controls=[name_field, unit_field, message],
                tight=True
            ),
            actions=[
                ft.Button(
                    content="Cancel",
                    on_click=lambda e: page.pop_dialog()
                ),
                ft.Button(
                    content="Save",
                    on_click=save
                )
            ]
        )
    )

def open_delete_item_dialog(
    page: ft.Page,
    item_id: int,
    name: str,
    items_list: ft.Column,
    inventory_item: ft.Dropdown,
    inventory_list: ft.Column
):
    record_count = count_inventory_for_item(item_id)

    if record_count:
        warning_text = (
            f'"{name}" has records. Are you sure you want to delete '
            "everything related to this item?"
        )
    else:
        warning_text = f'Delete "{name}"?'

    def confirm_delete(e):
        delete_item(item_id)

        load_items(page, items_list, inventory_item, inventory_list)
        load_inventory_items(page, inventory_item)
        load_inventory(page, inventory_list)

        page.pop_dialog()

    page.show_dialog(
        ft.AlertDialog(
            modal=True,
            title=ft.Text("Delete Item"),
            content=ft.Text(warning_text),
            actions=[
                ft.Button(
                    content="Cancel",
                    on_click=lambda e: page.pop_dialog()
                ),
                ft.Button(
                    content="Delete",
                    on_click=confirm_delete
                )
            ]
        )
    )


def load_inventory_items(page: ft.Page, inventory_item: ft.Dropdown):
    inventory_item.options.clear()

    for item_id, name, unit in get_items():
        inventory_item.options.append(
            ft.DropdownOption(
                key=str(item_id),
                text=name,
                data=unit
            )
        )

    page.update()

def select_inventory_item(
    e,
    inventory_item: ft.Dropdown,
    inventory_unit: ft.TextField
):
    if not e.control.value:
        inventory_unit.value = ""
        return

    selected_option = next(
        (
            option
            for option in inventory_item.options
            if option.key == e.control.value
        ),
        None
    )

    if selected_option:
        inventory_unit.value = selected_option.data

    inventory_unit.update()


def add_item(
    page: ft.Page,
    item_name: ft.TextField,
    item_unit: ft.TextField,
    items_list: ft.Column,
    message: ft.Text,
    inventory_item: ft.Dropdown,
    inventory_list: ft.Column
):
    name = item_name.value.strip()
    unit = item_unit.value.strip()

    if not name or not unit:
        return

    if not add_item_to_database(name, unit):
        message.value = "Item already exists."
        page.update()
        return

    message.value = ""
    item_name.value = ""
    item_unit.value = ""

    load_items(page, items_list, inventory_item, inventory_list)
    load_inventory_items(page, inventory_item)

    page.pop_dialog()


def load_inventory(page: ft.Page, inventory_list: ft.Column):
    inventory_list.controls.clear()

    rows = get_inventory()

    if not rows:
        inventory_list.controls.append(
            ft.Text(
                "No inventory recorded yet.",
                italic=True,
                color=COLOUR_TEXT
            )
        )

    for record_id, date, item_name, amount, unit in rows:
        inventory_list.controls.append(
            ft.Container(
                content=ft.Row(
                    controls=[
                        ft.Column(
                            controls=[
                                ft.Text(item_name, weight=ft.FontWeight.W_500),
                                ft.Text(
                                    f"{date} — {amount} {unit}",
                                    size=12,
                                    color= COLOUR_TEXT
                                )
                            ],
                            spacing=0,
                            expand=True
                        ),
                        ft.IconButton(
                            icon=ft.Icons.EDIT,
                            tooltip="Edit record",
                            icon_size=18,
                            on_click=lambda e, rid=record_id, dt=date, amt=amount: (
                                open_edit_inventory_dialog(
                                    page,
                                    rid,
                                    dt,
                                    amt,
                                    inventory_list
                                )
                            )
                        ),
                        ft.IconButton(
                            icon=ft.Icons.DELETE,
                            tooltip="Delete record",
                            icon_size=18,
                            on_click=lambda e, rid=record_id: (
                                open_delete_inventory_dialog(
                                    page,
                                    rid,
                                    inventory_list
                                )
                            )
                        )
                    ],
                    alignment=ft.MainAxisAlignment.SPACE_BETWEEN
                ),
                padding=ft.Padding.symmetric(vertical=2, horizontal=8),
                border_radius=8,
                border=ft.Border.only(
                    bottom=ft.BorderSide(1, COLOUR_ROW_SEPARATOR)
                )
            )
        )

    page.update()

def open_edit_inventory_dialog(
    page: ft.Page,
    record_id: int,
    date: str,
    amount: float,
    inventory_list: ft.Column
):
    amount_field = ft.TextField(label="Amount", value=str(amount))
    message = ft.Text()

    date_field = ft.TextField(
        label="Date",
        value=date,
        read_only=True,
        suffix_icon=ft.Icons.CALENDAR_MONTH,
        on_click=lambda e: page.show_dialog(date_picker)
    )

    def on_date_change(e):
        date_field.value = picked_date_str(e.control.value)
        date_field.update()

    date_picker = ft.DatePicker(
        value=datetime.strptime(date, "%Y-%m-%d"),
        first_date=datetime(2020, 1, 1),
        last_date=datetime(2100, 12, 31),
        on_change=on_date_change
    )

    def save(e):
        new_date = date_field.value.strip()
        new_amount_text = amount_field.value.strip()

        try:
            new_amount = float(new_amount_text)
        except ValueError:
            message.value = "Amount must be a number."
            page.update()
            return

        if new_amount < 0:
            message.value = "Amount cannot be negative."
            page.update()
            return

        if not update_inventory_record(record_id, new_date, new_amount):
            message.value = (
                "A record already exists for this item and date."
            )
            page.update()
            return

        load_inventory(page, inventory_list)
        page.pop_dialog()

    page.show_dialog(
        ft.AlertDialog(
            modal=True,
            title=ft.Text("Edit Inventory Record"),
            content=ft.Column(
                controls=[date_field, amount_field, message],
                tight=True
            ),
            actions=[
                ft.Button(
                    content="Cancel",
                    on_click=lambda e: page.pop_dialog()
                ),
                ft.Button(
                    content="Save",
                    on_click=save
                )
            ]
        )
    )

def open_delete_inventory_dialog(
    page: ft.Page,
    record_id: int,
    inventory_list: ft.Column
):
    def confirm_delete(e):
        delete_inventory_record(record_id)
        load_inventory(page, inventory_list)
        page.pop_dialog()

    page.show_dialog(
        ft.AlertDialog(
            modal=True,
            title=ft.Text("Delete Record"),
            content=ft.Text("Delete this inventory record?"),
            actions=[
                ft.Button(
                    content="Cancel",
                    on_click=lambda e: page.pop_dialog()
                ),
                ft.Button(
                    content="Delete",
                    on_click=confirm_delete
                )
            ]
        )
    )


def add_inventory_record(
    page: ft.Page,
    inventory_date: ft.TextField,
    inventory_item: ft.Dropdown,
    inventory_amount: ft.TextField,
    inventory_unit: ft.TextField,
    inventory_message: ft.Text,
    inventory_list: ft.Column
):
    date = inventory_date.value.strip()
    item_id = inventory_item.value
    amount = inventory_amount.value.strip()
    unit = inventory_unit.value.strip()

    if not date or not item_id or not amount or not unit:
        inventory_message.value = "Please fill in all fields."
        page.update()
        return

    try:
        date = datetime.strptime(
            date,
            "%Y-%m-%d"
        ).strftime("%Y-%m-%d")
    except ValueError:
        inventory_message.value = "Date must be YYYY-MM-DD."
        page.update()
        return

    try:
        amount = float(amount)
    except ValueError:
        inventory_message.value = "Amount must be a number."
        page.update()
        return

    if amount < 0:
        inventory_message.value = "Amount cannot be negative."
        page.update()
        return

    if not add_inventory(
        date,
        int(item_id),
        amount,
        unit
    ):
        inventory_message.value = (
            "An inventory record already exists "
            "for this item and date."
        )
        page.update()
        return

    inventory_message.value = ""
    inventory_amount.value = ""
    inventory_unit.value = ""

    load_inventory(page, inventory_list)

    page.pop_dialog()

def get_inventory_table_data(start_date=None, end_date=None, item_ids=None):
    items = get_items()

    if item_ids is not None:
        items = [item for item in items if item[0] in item_ids]

    if not items:
        return [], []

    item_names = {name for _, name, _ in items}

    inventory = get_inventory()
    inventory = [record for record in inventory if record[2] in item_names]

    if start_date:
        inventory = [record for record in inventory if record[1] >= start_date]

    if end_date:
        inventory = [record for record in inventory if record[1] <= end_date]

    # Create a lookup:
    # (date, item_id) -> amount
    inventory_lookup = {}

    for record_id, date, item_name, amount, unit in inventory:
        for item_id, name, item_unit in items:
            if name == item_name:
                inventory_lookup[(date, item_id)] = amount
                break

    # Determine the date range to render
    if start_date and end_date:
        range_start = datetime.strptime(start_date, "%Y-%m-%d")
        range_end = datetime.strptime(end_date, "%Y-%m-%d")
    elif inventory:
        dates = sorted(set(record[1] for record in inventory))
        range_start = datetime.strptime(dates[0], "%Y-%m-%d")
        range_end = datetime.strptime(dates[-1], "%Y-%m-%d")
    else:
        return items, []

    current_date = range_start
    all_dates = []

    while current_date <= range_end:
        all_dates.append(
            current_date.strftime("%Y-%m-%d")
        )

        current_date += timedelta(days=1)

    # Build the table rows
    rows = []

    for date in all_dates:
        row = [date]

        for item_id, name, unit in items:
            amount = inventory_lookup.get(
                (date, item_id)
            )

            if amount is None:
                row.append("No Data")
            else:
                row.append(str(amount))

        rows.append(row)

    return items, rows

def build_inventory_table(start_date=None, end_date=None, item_ids=None):
    items, rows = get_inventory_table_data(start_date, end_date, item_ids)

    columns = [
        fdt.DataColumn2(
            label=ft.Text("Date"),
            fixed_width=120
        )
    ]

    for item_id, name, unit in items:
        columns.append(
            fdt.DataColumn2(
                label=ft.Text(f"{name} ({unit})"),
                fixed_width=180
            )
        )

    table_rows = []

    for row in rows:
        cells = []

        for value in row:
            cells.append(
                ft.DataCell(
                    ft.Text(value)
                )
            )

        table_rows.append(
            ft.DataRow(
                cells=cells
            )
        )

    return fdt.DataTable2(
        columns=columns,
        rows=table_rows,
        fixed_top_rows=1,
        fixed_left_columns=1,
        min_width=1200,
        visible_horizontal_scroll_bar=True,
        visible_vertical_scroll_bar=True
    )

def show_inventory_table(page: ft.Page, inventory_table_dialog: ft.AlertDialog):
    filters = {"start_date": None, "end_date": None, "item_ids": None}

    save_message = ft.Text()

    table_container = ft.Container(
        content=build_inventory_table(),
        height=400,
        width=900
    )

    def rebuild_table():
        table_container.content = build_inventory_table(
            filters["start_date"],
            filters["end_date"],
            filters["item_ids"]
        )

    start_field = ft.TextField(
        label="From",
        width=150,
        read_only=True,
        suffix_icon=ft.Icons.CALENDAR_MONTH,
        on_click=lambda e: page.show_dialog(start_picker)
    )

    end_field = ft.TextField(
        label="To",
        width=150,
        read_only=True,
        suffix_icon=ft.Icons.CALENDAR_MONTH,
        on_click=lambda e: page.show_dialog(end_picker)
    )

    def on_start_change(e):
        start_field.value = picked_date_str(e.control.value)
        start_field.update()

    def on_end_change(e):
        end_field.value = picked_date_str(e.control.value)
        end_field.update()

    recorded_dates = [record[1] for record in get_inventory()]

    if recorded_dates:
        earliest_date = datetime.strptime(min(recorded_dates), "%Y-%m-%d")
        latest_date = datetime.strptime(max(recorded_dates), "%Y-%m-%d")
    else:
        earliest_date = datetime.now()
        latest_date = datetime.now()

    start_picker = ft.DatePicker(
        first_date=earliest_date,
        last_date=latest_date,
        on_change=on_start_change
    )

    end_picker = ft.DatePicker(
        first_date=earliest_date,
        last_date=latest_date,
        on_change=on_end_change
    )

    item_checkboxes = {}
    checkbox_controls = []

    for item_id, name, unit in get_items():
        checkbox = ft.Checkbox(label=f"{name} ({unit})", value=True)
        item_checkboxes[item_id] = checkbox
        checkbox_controls.append(checkbox)

    items_filter_column = ft.Column(
        controls=checkbox_controls,
        scroll=ft.ScrollMode.AUTO,
        height=100,
        spacing=0
    )

    def apply_filters(e):
        filters["start_date"] = start_field.value.strip() or None
        filters["end_date"] = end_field.value.strip() or None

        selected = [
            item_id
            for item_id, checkbox in item_checkboxes.items()
            if checkbox.value
        ]
        filters["item_ids"] = (
            selected if len(selected) < len(item_checkboxes) else None
        )

        rebuild_table()
        page.update()

    def clear_filters(e):
        start_field.value = ""
        end_field.value = ""

        for checkbox in item_checkboxes.values():
            checkbox.value = True

        filters["start_date"] = None
        filters["end_date"] = None
        filters["item_ids"] = None

        rebuild_table()
        page.update()

    export_csv_button = ft.Button(
        content="Export CSV",
        on_click=lambda e: save_inventory(
            page, save_message, "csv",
            filters["start_date"], filters["end_date"], filters["item_ids"]
        )
    )

    export_xlsx_button = ft.Button(
        content="Export XLSX",
        on_click=lambda e: save_inventory(
            page, save_message, "xlsx",
            filters["start_date"], filters["end_date"], filters["item_ids"]
        )
    )

    export_pdf_button = ft.Button(
        content="Export PDF",
        on_click=lambda e: save_inventory(
            page, save_message, "pdf",
            filters["start_date"], filters["end_date"], filters["item_ids"]
        )
    )

    close_button = ft.Button(
        content="Close",
        on_click=lambda e: page.pop_dialog()
    )

    inventory_table_dialog.content = ft.Container(
        content=ft.Column(
            controls=[
                ft.Row(
                    controls=[
                        start_field,
                        end_field,
                        ft.Button(
                            content="Apply Filters",
                            on_click=apply_filters
                        ),
                        ft.Button(
                            content="Clear Filters",
                            on_click=clear_filters
                        )
                    ],
                    spacing=10,
                    wrap=True
                ),
                ft.Text("Filter Items", size=12, color=COLOUR_TEXT),
                items_filter_column,
                ft.Divider(height=1, color=COLOUR_DIVIDER),
                table_container
            ],
            spacing=10,
            scroll=ft.ScrollMode.AUTO,
            expand=True
        ),
        height=600,
        width=900
    )

    inventory_table_dialog.actions = [
        ft.Row(
            controls=[
                save_message,
                export_csv_button,
                export_xlsx_button,
                export_pdf_button,
                close_button
            ],
            alignment=ft.MainAxisAlignment.END,
            spacing=10,
            run_spacing=10,
            wrap=True,
            width=860
        )
    ]

    page.show_dialog(inventory_table_dialog)

def get_inventory_table_headers(items):
    headers = ["Date"]

    for item_id, name, unit in items:
        headers.append(f"{name} ({unit})")

    return headers

def export_inventory_csv(filename, start_date=None, end_date=None, item_ids=None):
    items, rows = get_inventory_table_data(start_date, end_date, item_ids)
    headers = get_inventory_table_headers(items)

    with open(
        filename,
        "w",
        newline="",
        encoding="utf-8"
    ) as file:
        writer = csv.writer(file)

        writer.writerow(headers)

        for row in rows:
            writer.writerow(row)

def export_inventory_xlsx(filename, start_date=None, end_date=None, item_ids=None):
    items, rows = get_inventory_table_data(start_date, end_date, item_ids)
    headers = get_inventory_table_headers(items)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Inventory"

    sheet.append(headers)

    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            start_color=COLOUR_XLSX_CELL,
            end_color=COLOUR_XLSX_CELL,
            fill_type="solid"
        )

    for row in rows:
        sheet.append(row)

    for column_cells in sheet.columns:
        max_length = max(
            len(str(cell.value)) for cell in column_cells
        )
        sheet.column_dimensions[
            column_cells[0].column_letter
        ].width = min(max(max_length + 2, 10), 40)

    workbook.save(filename)

def export_inventory_pdf(filename, start_date=None, end_date=None, item_ids=None):
    items, rows = get_inventory_table_data(start_date, end_date, item_ids)
    headers = get_inventory_table_headers(items)

    styles = getSampleStyleSheet()
    title = Paragraph("Kitchen Inventory", styles["Title"])

    table = Table([headers] + rows, repeatRows=1)
    table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), pdf_colors.HexColor(COLOUR_PDF_HEADER_BG)),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, pdf_colors.grey),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ])
    )

    doc = SimpleDocTemplate(filename, pagesize=landscape(letter))
    doc.build([title, table])

EXPORTERS = {
    "csv": export_inventory_csv,
    "xlsx": export_inventory_xlsx,
    "pdf": export_inventory_pdf,
}

def save_inventory(
    page: ft.Page,
    save_message: ft.Text,
    file_format: str,
    start_date=None,
    end_date=None,
    item_ids=None
):
    items, rows = get_inventory_table_data(start_date, end_date, item_ids)

    if not rows:
        save_message.value = "No inventory data to export."
        save_message.color = COLOUR_SAVE_FAIL
        page.update()
        return

    try:
        downloads_dir = Path.home() / "Downloads"
        downloads_dir.mkdir(parents=True, exist_ok=True)

        filename = downloads_dir / (
            f"inventory_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            f".{file_format}"
        )

        EXPORTERS[file_format](str(filename), start_date, end_date, item_ids)

        save_message.value = f"Saved to {filename}"
        save_message.color = COLOUR_SAVE_SUCCESS

    except Exception as error:
        save_message.value = f"Save failed: {error}"
        save_message.color = COLOUR_SAVE_FAIL

    page.update()

def open_settings_dialog(page: ft.Page):
    theme_fields = {}

    def make_color_row(label, key):
        current_value = THEME[key]

        swatch = ft.Container(
            width=24,
            height=24,
            border_radius=4,
            bgcolor=(
                current_value
                if current_value.startswith("#")
                else f"#{current_value}"
            ),
            border=ft.Border.all(1, COLOUR_BORDER)
        )

        def on_change(e, sw=swatch):
            value = e.control.value.strip()
            if is_valid_hex_color(value):
                sw.bgcolor = value if value.startswith("#") else f"#{value}"
                sw.update()

        field = ft.TextField(
            label=label,
            value=current_value,
            width=160,
            on_change=on_change
        )
        theme_fields[key] = field

        return ft.Row(controls=[swatch, field], spacing=10)

    color_rows = [
        make_color_row("Text", "text"),
        make_color_row("Card Background", "card_bg"),
        make_color_row("Border", "border"),
        make_color_row("Divider", "divider"),
        make_color_row("Row Separator", "row_separator"),
        make_color_row("XLSX Header BG", "xlsx_header_bg"),
        make_color_row("PDF Header BG", "pdf_header_bg"),
        make_color_row("Save Success", "save_success"),
        make_color_row("Save Fail", "save_fail")
    ]

    width_field = ft.TextField(
        label="Window Width", value=str(WINDOW["width"]), width=160
    )
    height_field = ft.TextField(
        label="Window Height", value=str(WINDOW["height"]), width=160
    )
    min_width_field = ft.TextField(
        label="Min Width", value=str(WINDOW["min_width"]), width=160
    )
    min_height_field = ft.TextField(
        label="Min Height", value=str(WINDOW["min_height"]), width=160
    )

    message = ft.Text()

    def save(e):
        new_theme = {}

        for key, field in theme_fields.items():
            value = field.value.strip()

            if not is_valid_hex_color(value):
                message.value = f"Invalid color for {key}: {value}"
                message.color = COLOUR_SAVE_FAIL
                page.update()
                return

            new_theme[key] = value

        # openpyxl fill colors must not have a leading '#'
        new_theme["xlsx_header_bg"] = new_theme["xlsx_header_bg"].lstrip("#")

        try:
            new_window = {
                "width": int(width_field.value.strip()),
                "height": int(height_field.value.strip()),
                "min_width": int(min_width_field.value.strip()),
                "min_height": int(min_height_field.value.strip())
            }

            if any(value <= 0 for value in new_window.values()):
                raise ValueError

        except ValueError:
            message.value = "Window sizes must be positive whole numbers."
            message.color = COLOUR_SAVE_FAIL
            page.update()
            return

        save_config({"theme": new_theme, "window": new_window})

        message.value = "Saved — restart the app to apply changes."
        message.color = COLOUR_SAVE_SUCCESS
        page.update()

    page.show_dialog(
        ft.AlertDialog(
            modal=True,
            title=ft.Text("Settings"),
            content=ft.Column(
                controls=[
                    ft.Text(
                        "Theme Colors", size=14, weight=ft.FontWeight.BOLD
                    ),
                    *color_rows,
                    ft.Divider(height=1, color=COLOUR_DIVIDER),
                    ft.Text(
                        "Window Size", size=14, weight=ft.FontWeight.BOLD
                    ),
                    ft.Row(
                        controls=[width_field, height_field], spacing=10
                    ),
                    ft.Row(
                        controls=[min_width_field, min_height_field],
                        spacing=10
                    ),
                    message
                ],
                spacing=12,
                scroll=ft.ScrollMode.AUTO,
                width=380,
                height=500
            ),
            actions=[
                ft.Button(
                    content="Close",
                    on_click=lambda e: page.pop_dialog()
                ),
                ft.Button(
                    content="Save",
                    on_click=save
                )
            ],
            actions_alignment=ft.MainAxisAlignment.END
        )
    )

def main(page: ft.Page):
    page.title = "Kitchen Inventory"

    page.window.min_width = WINDOW["min_width"]
    page.window.min_height = WINDOW["min_height"]

    page.window.width = WINDOW["width"]
    page.window.height = WINDOW["height"]
    page.update()

    items_list = ft.Column(
        scroll=ft.ScrollMode.AUTO,
        expand=True
    )
    inventory_list = ft.Column(
        scroll=ft.ScrollMode.AUTO,
        expand=True
    )

    message = ft.Text()
    inventory_message = ft.Text()

    # -------------------------
    # Add Item controls
    # -------------------------

    item_name = ft.TextField(
        label="Item name",
        width=320,
        on_submit=lambda: add_item(
            page,
            item_name,
            item_unit,
            items_list,
            message,
            inventory_item,
            inventory_list
        )
    )

    item_unit = ft.TextField(
        label="Unit",
        width=320,
        on_submit=lambda e: add_item(
            page,
            item_name,
            item_unit,
            items_list,
            message,
            inventory_item,
            inventory_list
        )
    )

    add_button = ft.Button(
        content="Add",
        on_click=lambda: add_item(
            page,
            item_name,
            item_unit,
            items_list,
            message,
            inventory_item,
            inventory_list
        )
    )

    close_add_item_button = ft.Button(
        content="Cancel",
        on_click=lambda: page.pop_dialog()
    )

    add_item_dialog = ft.AlertDialog(
        modal=True,
        title=ft.Text("Add Item"),
        content=ft.Column(
            controls=[
                item_name,
                item_unit,
                message
            ],
            tight=True,
            spacing=15,
            width=320
        ),
        actions=[
            close_add_item_button,
            add_button
        ],
        actions_alignment=ft.MainAxisAlignment.END
    )

    open_add_item_button = ft.Button(
        content="Add Item",
        icon=ft.Icons.ADD,
        on_click=lambda: page.show_dialog(add_item_dialog)
    )

    # -------------------------
    # Inventory controls
    # -------------------------

    inventory_item = ft.Dropdown(
        label="Item",
        width=320,
        on_select=lambda e: select_inventory_item(
            e,
            inventory_item,
            inventory_unit
        )
    )

    inventory_date = ft.TextField(
        label="Date",
        width=320,
        value=datetime.now().strftime("%Y-%m-%d"),
        read_only=True,
        suffix_icon=ft.Icons.CALENDAR_MONTH,
        on_click=lambda e: page.show_dialog(inventory_date_picker)
    )

    def on_inventory_date_change(e):
        inventory_date.value = picked_date_str(e.control.value)
        inventory_date.update()

    inventory_date_picker = ft.DatePicker(
        value=datetime.now(),
        first_date=datetime(2020, 1, 1),
        last_date=datetime(2100, 12, 31),
        on_change=on_inventory_date_change
    )

    inventory_amount = ft.TextField(
        label="Amount",
        width=320,
        on_submit=lambda e: add_inventory_record(
            page,
            inventory_date,
            inventory_item,
            inventory_amount,
            inventory_unit,
            inventory_message,
            inventory_list
        )
    )

    inventory_unit = ft.TextField(
        label="Unit",
        width=320,
        read_only=True,
        on_submit=lambda e: add_inventory_record(
            page,
            inventory_date,
            inventory_item,
            inventory_amount,
            inventory_unit,
            inventory_message,
            inventory_list
        )
    )

    inventory_add_button = ft.Button(
        content="Add Inventory",
        on_click=lambda e: add_inventory_record(
            page,
            inventory_date,
            inventory_item,
            inventory_amount,
            inventory_unit,
            inventory_message,
            inventory_list
        )
    )

    close_inventory_button = ft.Button(
        content="Cancel",
        on_click=lambda: page.pop_dialog()
    )

    inventory_dialog = ft.AlertDialog(
        modal=True,
        title=ft.Text("Inventory an Item"),
        content=ft.Column(
            controls=[
                inventory_date,
                inventory_item,
                inventory_amount,
                inventory_unit,
                inventory_message
            ],
            tight=True,
            spacing=15,
            width=320
        ),
        actions=[
            close_inventory_button,
            inventory_add_button
        ],
        actions_alignment=ft.MainAxisAlignment.END
    )

    open_inventory_button = ft.Button(
        content="Inventory an Item",
        icon=ft.Icons.PLAYLIST_ADD_CHECK,
        on_click=lambda: page.show_dialog(inventory_dialog)
    )

    inventory_table_dialog = ft.AlertDialog(
        modal=True,
        title=ft.Text("Current Inventory")
    )

    open_inventory_table_button = ft.Button(
        content="Current Inventory",
        icon=ft.Icons.TABLE_CHART,
        on_click=lambda e: show_inventory_table(
            page,
            inventory_table_dialog
        )
    )

    # -------------------------
    # Main page
    # -------------------------

    page.padding = 24

    items_card = ft.Container(
        content=ft.Column(
            controls=[
                ft.Row(
                    controls=[
                        ft.Icon(ft.Icons.SHOPPING_BASKET_OUTLINED, size=20),
                        ft.Text("Items", size=18, weight=ft.FontWeight.BOLD)
                    ],
                    spacing=8
                ),
                ft.Divider(height=1, color=COLOUR_DIVIDER),
                items_list
            ],
            spacing=12,
            expand=True
        ),
        padding=16,
        border_radius=12,
        border=ft.Border.all(1, COLOUR_BORDER),
        bgcolor=COLOUR_CARD_BG,
        expand=True
    )

    inventory_card = ft.Container(
        content=ft.Column(
            controls=[
                ft.Row(
                    controls=[
                        ft.Icon(ft.Icons.HISTORY, size=20),
                        ft.Text(
                            "Inventory History",
                            size=18,
                            weight=ft.FontWeight.BOLD
                        )
                    ],
                    spacing=8
                ),
                ft.Divider(height=1, color=COLOUR_DIVIDER),
                inventory_list
            ],
            spacing=12,
            expand=True
        ),
        padding=16,
        border_radius=12,
        border=ft.Border.all(1, COLOUR_BORDER),
        bgcolor=COLOUR_CARD_BG,
        expand=True
    )

    page.add(
        ft.Column(
            controls=[
                ft.Row(
                    controls=[
                        ft.Icon(ft.Icons.KITCHEN, size=32),
                        ft.Column(
                            controls=[
                                ft.Text(
                                    "Kitchen Inventory",
                                    size=26,
                                    weight=ft.FontWeight.BOLD
                                ),
                                ft.Text(
                                    "Track what's in stock and keep a "
                                    "history over time",
                                    size=13,
                                    color=COLOUR_TEXT
                                )
                            ],
                            spacing=0
                        )
                    ],
                    spacing=12
                ),

                ft.Row(
                    controls=[
                        open_add_item_button,
                        open_inventory_button,
                        open_inventory_table_button,
                        ft.Button(
                            content="Settings",
                            icon=ft.Icons.SETTINGS,
                            on_click=lambda e: open_settings_dialog(page)
                        )
                    ],
                    spacing=12
                ),

                ft.Row(
                    controls=[items_card, inventory_card],
                    spacing=20,
                    expand=True
                )
            ],
            spacing=20,
            expand=True
        )
    )

    load_items(page, items_list, inventory_item, inventory_list)
    load_inventory_items(page, inventory_item)
    load_inventory(page, inventory_list)


ft.run(main)